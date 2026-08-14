"""File import service: container parsing, preview and atomic apply.

Python preserves raw CSV/XLSX cell values, supplied state, mappings and source
coordinates. The Go import preview service and FieldValueKernel exclusively own
target-field conversion, defaults, blank semantics and constraints.

Process
----
* :meth:`preview` reads the granted file (streaming for large workbooks),
  auto-maps or applies the explicit column mapping, resolves explicit relation
  lookups, delegates raw cells to the authoritative Go preview, and returns an
  :class:`ImportPlan` bound to the source hash + capability hash via a
  single-use token.
* :meth:`apply` submits every valid planned row in one frozen mutation request.
  Cancellation is checked before submission; a rejected request therefore
  leaves zero rows committed.

The Qt/controller ``confirm_cb`` callback is gone: preview is zero-write and
returns the full plan; the host shows it and the user confirms via apply.
"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import time
import uuid
from collections.abc import Awaitable, Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from backend.application.paste_service import PasteMutationPort
from backend.contracts.data_io import (
    MAX_ATOMIC_IMPORT_ROWS,
    ApplyImportParams,
    ApplyImportResult,
    ImportCellDiagnostic,
    ImportChunkResult,
    ImportColumnMapping,
    ImportPlan,
    ImportPlanRow,
    ImportPreviewToken,
    ImportRelationResolution,
    ImportSummary,
    PreviewImportParams,
)
from backend.contracts.data_profile import CollectionProfile
from backend.contracts.paste import PastePlanRow

#: How long an import preview token remains valid (seconds).
IMPORT_TOKEN_TTL_SECONDS: float = 10 * 60.0

#: Compatibility default retained by the public contract. Apply is atomic.
DEFAULT_CHUNK_SIZE: int = 500


class ImportFlowError(Exception):
    """An import error carrying an RPC-friendly ``code``."""

    def __init__(self, message: str, *, code: str, data: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.data = data

    @property
    def rpc_error_data(self) -> dict[str, Any]:
        exposed: dict[str, Any] = {"code": self.code}
        if self.data:
            exposed.update(self.data)
        return exposed


@dataclass(frozen=True)
class RelationImportTarget:
    """Live-schema proof for one explicit relation import mapping."""

    relation_id: str
    target_field: str
    target_collection: str
    target_primary_key: str
    match_field: str


@dataclass(frozen=True)
class RelationImportBatchResult:
    """Server-confirmed result of one atomic relation-aware import chunk."""

    created_row_keys: list[str]
    updated_row_keys: list[str]
    request_id: str = ""


class RelationImportProvider(Protocol):
    """Permission-scoped adapter for relation resolution and atomic apply.

    The implementation uses the current product session. ``inspect_mapping``
    must reject non-PK/non-unique match
    fields. ``apply_chunk`` owns one transaction containing any requested target
    creation and the source-row mutation, and deduplicates by ``idempotency_key``.
    """

    async def inspect_mapping(
        self,
        *,
        collection: str,
        target_field: str,
        relation_id: str,
        match_field: str,
    ) -> RelationImportTarget: ...

    async def find_exact(self, target: RelationImportTarget, value: Any) -> list[Any]: ...

    async def apply_chunk(
        self,
        *,
        collection: str,
        profile: CollectionProfile,
        rows: list[ImportPlanRow],
        mode: str,
        upsert_key: str | None,
        idempotency_key: str,
    ) -> RelationImportBatchResult: ...


class ImportMutationPort(PasteMutationPort, Protocol):
    """Mutation port with Go-owned raw-cell normalization for import preview."""

    async def preview_import(
        self,
        *,
        collection: str,
        schema_revision: str,
        rows: list[dict[str, Any]],
        row_modes: list[str] | None = None,
    ) -> dict[str, Any]: ...


# ---------------------------------------------------------------------------
# File reading (streaming, cancellation-aware)
# ---------------------------------------------------------------------------


class SourceFile:
    """A granted import source file, read lazily.

    Excel workbooks are read sheet-by-sheet; CSV files are read row-by-row. The
    reader never loads the entire file into memory at once.
    """

    def __init__(self, path: str) -> None:
        self._path = Path(path)

    def read_header_and_rows(
        self,
        *,
        max_rows: int = MAX_ATOMIC_IMPORT_ROWS,
        sheet: str | None = None,
    ) -> tuple[list[str], list[list[Any]], str]:
        """Read the header row + data rows.

        Returns ``(header, rows, source_hash)``. For XLSX the first sheet (or the
        named ``sheet``) is used; for CSV the single stream is used. Files with
        more than ``max_rows`` data rows are rejected instead of truncated,
        because apply is one atomic mutation with the same fixed row limit.
        ``source_hash`` is the SHA-256 of the file bytes (binds the preview).
        """
        source_hash = hashlib.sha256(self._path.read_bytes()).hexdigest()
        suffix = self._path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            header, rows = self._read_xlsx(max_rows=max_rows, sheet=sheet)
            return header, rows, source_hash
        if suffix == ".csv":
            header, rows = self._read_csv(max_rows=max_rows)
            return header, rows, source_hash
        raise ImportFlowError(
            f"unsupported file type {suffix!r}",
            code="import_unsupported_format",
        )

    def _read_xlsx(self, *, max_rows: int, sheet: str | None) -> tuple[list[str], list[list[Any]]]:
        from openpyxl import load_workbook

        wb = load_workbook(self._path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            if ws is None:
                raise ImportFlowError(
                    "workbook has no readable worksheet",
                    code="import_empty_workbook",
                )
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(c) if c is not None else "" for c in next(rows_iter)]
            except StopIteration:
                return [], []
            try:
                data: list[list[Any]] = []
                for row in rows_iter:
                    if len(data) >= max_rows:
                        raise _import_row_limit(max_rows)
                    data.append(list(row))
                return header, data
            finally:
                rows_iter.close()
        finally:
            wb.close()

    def _read_csv(self, *, max_rows: int) -> tuple[list[str], list[list[Any]]]:
        with self._path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            try:
                header = [str(c) for c in next(reader)]
            except StopIteration:
                return [], []
            data: list[list[Any]] = []
            for row in reader:
                if len(data) >= max_rows:
                    raise _import_row_limit(max_rows)
                data.append(list(row))
        return header, data


def _import_row_limit(max_rows: int) -> ImportFlowError:
    return ImportFlowError(
        f"import contains more than {max_rows} data rows",
        code="import_row_limit",
        data={"maxRows": max_rows},
    )


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------


def auto_map_columns(
    header: list[str],
    profile: CollectionProfile,
    explicit: list[ImportColumnMapping],
) -> tuple[dict[int, str], list[str]]:
    """Map source column indices → collection field names.

    Explicit mappings take precedence; remaining columns are auto-matched by
    field key (case-insensitive). Returns ``(mapping, unmatched_columns)``.
    """
    explicit_by_source = {m.source_column: m.target_field for m in explicit}
    create_fields = set(profile.create_fields)
    mapping: dict[int, str] = {}
    matched_sources: set[str] = set()
    for index, name in enumerate(header):
        clean = name.strip()
        if clean in explicit_by_source:
            target = explicit_by_source[clean]
            if target in create_fields:
                mapping[index] = target
                matched_sources.add(clean)
            continue
        lowered = clean.lower().replace(" ", "_")
        for field in profile.create_fields:
            if field.lower() == lowered:
                mapping[index] = field
                matched_sources.add(clean)
                break
    unmatched = [name for name in header if name.strip() not in matched_sources]
    return mapping, unmatched


# ---------------------------------------------------------------------------
# Import service
# ---------------------------------------------------------------------------


class _StoredImportPlan:
    """A preview plan retained server-side for a single apply."""

    __slots__ = (
        "capability_hash",
        "collection",
        "consumed",
        "expires_at",
        "idempotency_prefix",
        "mode",
        "rows",
        "schema_revision",
        "source_hash",
        "upsert_key",
    )

    def __init__(
        self,
        *,
        collection: str,
        schema_revision: str,
        capability_hash: str,
        source_hash: str,
        rows: list[ImportPlanRow],
        mode: str,
        upsert_key: str | None,
        expires_at: float,
    ) -> None:
        self.collection = collection
        self.schema_revision = schema_revision
        self.capability_hash = capability_hash
        self.source_hash = source_hash
        self.rows = rows
        self.mode = mode
        self.upsert_key = upsert_key
        self.expires_at = expires_at
        self.consumed = False
        self.idempotency_prefix: str | None = None


class ImportService:
    """Import preview and atomic apply over product-owned ports."""

    def __init__(
        self,
        *,
        client: Any,
        auth: Any,
        bulk: ImportMutationPort,
        profiles: dict[str, CollectionProfile],
        resolve_path: Callable[..., str],
        consume_grant: Callable[[str], None],
        relation_provider: RelationImportProvider | None = None,
        clock: Callable[[], float] = time.time,
    ) -> None:
        self._client = client
        self._auth = auth
        self._bulk = bulk
        self._profiles = profiles
        self._resolve_path = resolve_path
        self._consume_grant = consume_grant
        self._relation_provider = relation_provider
        self._clock = clock
        self._plans: dict[str, _StoredImportPlan] = {}
        self._apply_lock = asyncio.Lock()

    async def preview(self, params: PreviewImportParams) -> ImportPlan:
        profile = self._profile(params.collection)
        if params.schema_revision != profile.capability_hash:
            raise ImportFlowError(
                "schema changed since the grid was rendered",
                code="schema_mismatch",
                data={
                    "currentSchemaRevision": profile.capability_hash,
                    "expectedSchemaRevision": params.schema_revision,
                },
            )
        path = self._resolve_path(
            params.grant_id,
            purpose="import_source",
            direction="read",
        )
        source = SourceFile(path)
        header, rows, source_hash = source.read_header_and_rows()
        mapping, unmatched = auto_map_columns(header, profile, params.column_mapping)
        relations = {r.field: r for r in profile.relations}
        explicit_by_source = {item.source_column.strip(): item for item in params.column_mapping}
        relation_targets: dict[int, RelationImportTarget] = {}
        relation_mapping_errors: dict[int, tuple[str, str]] = {}
        for col_index, field in mapping.items():
            relation = relations.get(field)
            explicit = explicit_by_source.get(header[col_index].strip())
            if relation is None:
                if explicit and (explicit.relation_id or explicit.match_field):
                    relation_mapping_errors[col_index] = (
                        "relation_mapping_not_relation",
                        f"field {field!r} is not a relation",
                    )
                continue
            if explicit is None or explicit.relation_id is None or explicit.match_field is None:
                relation_mapping_errors[col_index] = (
                    "relation_mapping_required",
                    "relation columns require explicit relationId and matchField",
                )
                continue
            if relation.relation_id is not None and relation.relation_id != explicit.relation_id:
                relation_mapping_errors[col_index] = (
                    "relation_id_mismatch",
                    "relationId does not identify the mapped target field",
                )
                continue
            if self._relation_provider is None:
                relation_mapping_errors[col_index] = (
                    "relation_provider_unavailable",
                    "relation import resolution is not configured",
                )
                continue
            try:
                relation_targets[col_index] = await self._relation_provider.inspect_mapping(
                    collection=params.collection,
                    target_field=field,
                    relation_id=explicit.relation_id,
                    match_field=explicit.match_field,
                )
            except Exception as exc:
                relation_mapping_errors[col_index] = (
                    getattr(exc, "code", "relation_mapping_invalid"),
                    str(exc),
                )
        plan_rows: list[ImportPlanRow] = []
        for row_offset, raw_row in enumerate(rows):
            source_row = row_offset + 2  # 1-based + header
            values: dict[str, Any] = {}
            diagnostics: list[ImportCellDiagnostic] = []
            relation_resolutions: list[ImportRelationResolution] = []
            for col_index, field in mapping.items():
                raw = raw_row[col_index] if col_index < len(raw_row) else None
                relation = relations.get(field)
                mapping_error = relation_mapping_errors.get(col_index)
                if mapping_error is not None and raw is not None and raw != "":
                    diagnostics.append(
                        ImportCellDiagnostic(
                            row=source_row,
                            column=col_index + 1,
                            severity="error",
                            code=mapping_error[0],
                            message=mapping_error[1],
                            original_value=str(raw),
                        )
                    )
                    continue
                if relation is not None:
                    if raw is None or raw == "":
                        values[field] = None
                        continue
                    target = relation_targets[col_index]
                    assert self._relation_provider is not None
                    try:
                        matches = await self._relation_provider.find_exact(target, raw)
                    except Exception as exc:
                        diagnostics.append(
                            ImportCellDiagnostic(
                                row=source_row,
                                column=col_index + 1,
                                severity="error",
                                code=getattr(exc, "code", "relation_lookup_failed"),
                                message=str(exc),
                                original_value=str(raw),
                            )
                        )
                        continue
                    if len(matches) == 1:
                        values[field] = matches[0]
                        relation_resolutions.append(
                            ImportRelationResolution(
                                target_field=field,
                                relation_id=target.relation_id,
                                match_field=target.match_field,
                                source_value=raw,
                                state="matched",
                                matched_primary_key=matches[0],
                            )
                        )
                    elif len(matches) > 1:
                        diagnostics.append(
                            ImportCellDiagnostic(
                                row=source_row,
                                column=col_index + 1,
                                severity="error",
                                code="relation_match_ambiguous",
                                message=(
                                    f"exact match returned {len(matches)} target records; "
                                    "matchField must resolve to one record"
                                ),
                                original_value=str(raw),
                            )
                        )
                    else:
                        diagnostics.append(
                            ImportCellDiagnostic(
                                row=source_row,
                                column=col_index + 1,
                                severity="error",
                                code="relation_match_not_found",
                                message="exact match returned no target record",
                                original_value=str(raw),
                            )
                        )
                    continue
                # Preserve the exact parsed container value and key presence.
                # Target-field conversion and validation are owned by the Go
                # import preview service and FieldValueKernel.
                values[field] = raw
            plan_rows.append(
                ImportPlanRow(
                    source_row=source_row,
                    values=values,
                    diagnostics=diagnostics,
                    relation_resolutions=relation_resolutions,
                )
            )
        authoritative = await self._bulk.preview_import(
            collection=params.collection,
            schema_revision=params.schema_revision,
            rows=[row.values for row in plan_rows],
        )
        normalized_rows = authoritative.get("rows")
        if not isinstance(normalized_rows, list) or len(normalized_rows) != len(plan_rows):
            raise ImportFlowError(
                "invalid authoritative import preview",
                code="import_preview_invalid",
            )
        source_column_by_field = {field: index for index, field in mapping.items()}
        for row_index, authoritative_row in enumerate(normalized_rows):
            if not isinstance(authoritative_row, dict):
                raise ImportFlowError(
                    "invalid authoritative import preview row",
                    code="import_preview_invalid",
                )
            normalized_values = authoritative_row.get("values")
            raw_diagnostics = authoritative_row.get("diagnostics")
            if not isinstance(normalized_values, dict) or not isinstance(raw_diagnostics, list):
                raise ImportFlowError(
                    "invalid authoritative import preview row",
                    code="import_preview_invalid",
                )
            current = plan_rows[row_index]
            current.values = normalized_values
            for diagnostic in raw_diagnostics:
                if not isinstance(diagnostic, dict):
                    raise ImportFlowError(
                        "invalid authoritative import diagnostic",
                        code="import_preview_invalid",
                    )
                raw_field = diagnostic.get("field")
                field = raw_field if isinstance(raw_field, str) else ""
                column_index = source_column_by_field.get(field, 0)
                raw_row = rows[row_index]
                original = raw_row[column_index] if column_index < len(raw_row) else None
                current.diagnostics.append(
                    ImportCellDiagnostic(
                        row=current.source_row,
                        column=column_index + 1,
                        severity="error",
                        code=str(diagnostic.get("code") or "field.value.invalid"),
                        message=str(diagnostic.get("message") or "invalid field value"),
                        original_value="" if original is None else str(original),
                    )
                )
        error_rows = sum(
            any(item.severity == "error" for item in row.diagnostics) for row in plan_rows
        )
        warning_rows = sum(
            any(item.severity == "warning" for item in row.diagnostics) for row in plan_rows
        )
        error_count = sum(item.severity == "error" for row in plan_rows for item in row.diagnostics)
        warning_count = sum(
            item.severity == "warning" for row in plan_rows for item in row.diagnostics
        )
        summary = ImportSummary(
            total_rows=len(rows),
            valid_rows=len(rows) - error_rows,
            error_rows=error_rows,
            warning_rows=warning_rows,
            error_count=error_count,
            warning_count=warning_count,
        )
        token = self._mint_plan(
            collection=params.collection,
            schema_revision=params.schema_revision,
            capability_hash=profile.capability_hash,
            source_hash=source_hash,
            rows=plan_rows,
            mode=params.mode,
            upsert_key=params.upsert_key,
        )
        return ImportPlan(
            collection=params.collection,
            schema_revision=params.schema_revision,
            capability_hash=profile.capability_hash,
            source_hash=source_hash,
            summary=summary,
            rows=plan_rows,
            unmatched_columns=unmatched,
            token=token,
        )

    async def apply(
        self,
        params: ApplyImportParams,
        *,
        progress: Callable[[int, int, str], Awaitable[None]] | None = None,
        cancelled: Callable[[], bool] | None = None,
    ) -> ApplyImportResult:
        async with self._apply_lock:
            return await self._apply_serialized(
                params,
                progress=progress,
                cancelled=cancelled,
            )

    async def _apply_serialized(
        self,
        params: ApplyImportParams,
        *,
        progress: Callable[[int, int, str], Awaitable[None]] | None = None,
        cancelled: Callable[[], bool] | None = None,
    ) -> ApplyImportResult:
        profile = self._profile(params.collection)
        stored = self._plans.get(params.token)
        if stored is None:
            raise ImportFlowError("import token not found", code="import_token_unknown")
        if self._clock() >= stored.expires_at:
            raise ImportFlowError("import token expired", code="import_token_expired")
        if stored.consumed:
            raise ImportFlowError("import token already used", code="import_token_consumed")
        if stored.capability_hash != profile.capability_hash:
            raise ImportFlowError("schema changed since preview", code="schema_mismatch")
        valid_rows = [
            r for r in stored.rows if not any(d.severity == "error" for d in r.diagnostics)
        ]
        total = len(valid_rows)
        if cancelled and cancelled():
            raise asyncio.CancelledError
        requested_prefix = params.idempotency_prefix or (
            "imp-" + hashlib.sha256(params.token.encode("utf-8")).hexdigest()[:16]
        )
        if stored.idempotency_prefix is None:
            stored.idempotency_prefix = requested_prefix
        elif stored.idempotency_prefix != requested_prefix:
            raise ImportFlowError(
                "import token is bound to a different idempotency prefix",
                code="import_idempotency_mismatch",
            )
        prefix = stored.idempotency_prefix
        idempotency_key = f"{prefix}-0"
        bulk_rows = [
            PastePlanRow(
                kind="insert",
                changes={
                    field: {"before": None, "after": value} for field, value in row.values.items()
                },
            )
            for row in valid_rows
        ]
        requires_cross_table = stored.mode == "upsert" or any(
            resolution.state == "create"
            for row in valid_rows
            for resolution in row.relation_resolutions
        )
        try:
            if requires_cross_table:
                if self._relation_provider is None:
                    raise ImportFlowError(
                        "atomic relation/upsert import is not configured",
                        code="relation_provider_unavailable",
                    )
                relation_result = await self._relation_provider.apply_chunk(
                    collection=params.collection,
                    profile=profile,
                    rows=valid_rows,
                    mode=stored.mode,
                    upsert_key=stored.upsert_key,
                    idempotency_key=idempotency_key,
                )
                created_keys = relation_result.created_row_keys
                updated_keys = relation_result.updated_row_keys
                request_id = relation_result.request_id
            else:
                result = await self._bulk.apply(
                    collection=params.collection,
                    profile=profile,
                    rows=bulk_rows,
                    row_revisions={},
                    idempotency_key=idempotency_key,
                    schema_revision=stored.schema_revision,
                )
                if result.outcome == "pending":
                    raise ImportFlowError(
                        "import outcome is pending; retry with the same token",
                        code="import_pending",
                    )
                if result.outcome != "committed":
                    raise ImportFlowError(
                        "import conflicted; preview again",
                        code="import_conflict",
                    )
                created_keys = [str(key) for key in result.created_row_keys]
                updated_keys = [str(key) for key in result.updated_row_keys]
                request_id = result.request_id
        except Exception as exc:
            if progress:
                safe_code = getattr(exc, "code", exc.__class__.__name__)
                safe_parts = [f"atomic import failed [{safe_code}]"]
                cause = exc.__cause__
                safe_path = getattr(cause, "path", None)
                if isinstance(safe_path, str) and safe_path:
                    safe_parts.append(f"at {safe_path}")
                    safe_message = str(cause)
                    if safe_message:
                        safe_parts.append(safe_message)
                await progress(total, total, ": ".join(safe_parts))
            return ApplyImportResult(
                collection=params.collection,
                created_count=0,
                updated_count=0,
                failed_rows=[row.source_row for row in valid_rows],
            )
        chunk = ImportChunkResult(
            chunk_index=0,
            created_row_keys=created_keys,
            updated_row_keys=updated_keys,
            failed_rows=[],
            idempotency_key=idempotency_key,
        )
        if progress:
            await progress(total, total, "atomic import committed")
        stored.consumed = True
        self._consume_grant(params.grant_id)
        return ApplyImportResult(
            collection=params.collection,
            created_count=len(created_keys),
            updated_count=len(updated_keys),
            failed_rows=[],
            chunks=[chunk],
            request_ids=[request_id] if request_id else [],
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _profile(self, collection: str) -> CollectionProfile:
        profile = self._profiles.get(collection)
        if profile is None:
            raise ImportFlowError(
                f"collection {collection!r} is not in the product schema",
                code="schema_unknown",
            )
        return profile

    def _mint_plan(
        self,
        *,
        collection: str,
        schema_revision: str,
        capability_hash: str,
        source_hash: str,
        rows: list[ImportPlanRow],
        mode: str,
        upsert_key: str | None,
    ) -> ImportPreviewToken:
        token = f"imp-{uuid.uuid4().hex[:24]}"
        self._plans[token] = _StoredImportPlan(
            collection=collection,
            schema_revision=schema_revision,
            capability_hash=capability_hash,
            source_hash=source_hash,
            rows=rows,
            mode=mode,
            upsert_key=upsert_key,
            expires_at=self._clock() + IMPORT_TOKEN_TTL_SECONDS,
        )
        return ImportPreviewToken(
            token=token, expires_at=self._clock() + IMPORT_TOKEN_TTL_SECONDS, consumed=False
        )


__all__ = [
    "DEFAULT_CHUNK_SIZE",
    "IMPORT_TOKEN_TTL_SECONDS",
    "MAX_ATOMIC_IMPORT_ROWS",
    "ImportFlowError",
    "ImportService",
    "RelationImportBatchResult",
    "RelationImportProvider",
    "RelationImportTarget",
    "SourceFile",
    "auto_map_columns",
]

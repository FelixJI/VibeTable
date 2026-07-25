"""Export service: stream a full product query result to CSV/XLSX.

Covers the C1 Task 4 requirements:

* Reuses the product query contract (``filter/sort/search/fields``) and covers
  all matching rows via stable paging.
* Streams rows to CSV/XLSX so memory stays roughly constant regardless of result
  size. Reports progress via the task runtime.
* Honours field/row permissions through the product query port.
* Templates: :meth:`generate_template` writes a schema-derived template (column
  names, required hints, enum/relation notes) to an export-target grant.

Cancellation is cooperative between pages. Output is written to a sibling
temporary file and atomically replaces the selected target only on success;
cancelled/failed exports never expose a partial artifact.
"""

from __future__ import annotations

import asyncio
import csv
import json
import os
import uuid
from collections.abc import Awaitable, Callable
from contextlib import ExitStack
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from backend.contracts.data_io import ExportParams, ExportResult, TemplateResult
from backend.contracts.data_profile import CollectionProfile
from backend.contracts.query import TableQuery

#: Page size for the export paging loop.
EXPORT_PAGE_SIZE: int = 100


class ExportError(Exception):
    """An export error carrying an RPC-friendly ``code``."""

    def __init__(self, message: str, *, code: str) -> None:
        super().__init__(message)
        self.code = code

    @property
    def rpc_error_data(self) -> dict[str, Any]:
        return {"code": self.code}


@dataclass(frozen=True)
class AuthoritativeLookupColumn:
    lookup_id: str
    field_key: str


@dataclass(frozen=True)
class AuthoritativeLookupExportPage:
    """One full-dataset page returned by the Lookup query data plane."""

    rows: list[dict[str, Any]]
    columns: list[AuthoritativeLookupColumn]
    filtered_rows: int
    lookup_revision: str


class AuthoritativeLookupExportProvider(Protocol):
    """Adapter over the authoritative Lookup query endpoint.

    Implementations must execute the supplied full query under the current
    Product-session accountability. They must not derive Lookup cells from rows already
    loaded in the Web grid.
    """

    async def query_page(
        self,
        *,
        collection: str,
        fields: list[str],
        lookup_ids: list[str],
        lookup_revision: str,
        query: dict[str, Any],
        offset: int,
        limit: int,
    ) -> AuthoritativeLookupExportPage: ...


class QueryPage(Protocol):
    rows: list[dict[str, Any]]
    filtered_rows: int
    total_rows: int


class QueryPagePort(Protocol):
    async def query_page(
        self,
        *,
        table_id: str,
        query: dict[str, Any],
    ) -> QueryPage: ...


class ExportService:
    """C1 query-based export + template generation."""

    def __init__(
        self,
        *,
        query_port: QueryPagePort,
        profiles: dict[str, CollectionProfile],
        resolve_path: Callable[..., str],
        lookup_provider: AuthoritativeLookupExportProvider | None = None,
    ) -> None:
        self._query_port = query_port
        self._profiles = profiles
        self._resolve_path = resolve_path
        self._lookup_provider = lookup_provider

    async def export(
        self,
        params: ExportParams,
        *,
        progress: Callable[[int, int, str], Awaitable[None]] | None = None,
        cancelled: Callable[[], bool] | None = None,
    ) -> ExportResult:
        profile = self._profile(params.collection)
        path = self._resolve_path(
            params.grant_id,
            purpose="export_target",
            direction="write",
        )
        output_columns = list(profile.fields)
        if params.include_relations:
            for relation in profile.relations:
                for display in relation.display_fields:
                    col = f"{relation.field}.{display}"
                    if col not in output_columns:
                        output_columns.append(col)
        target = Path(path)
        temporary = target.with_name(f".{target.name}.vibetable-{uuid.uuid4().hex}.tmp")
        rows_written = 0
        fmt = params.format
        try:
            if params.lookup_ids:
                if self._lookup_provider is None:
                    raise ExportError(
                        "authoritative Lookup export is not configured",
                        code="lookup_export_provider_missing",
                    )
                assert params.lookup_revision is not None
                rows_written = await self._export_with_lookups(
                    str(temporary),
                    profile,
                    params,
                    output_columns,
                    progress,
                    cancelled,
                )
            else:
                query = TableQuery.model_validate(params.query)
                if fmt == "csv":
                    rows_written = await self._export_csv(
                        str(temporary),
                        profile,
                        query,
                        output_columns,
                        progress,
                        cancelled,
                    )
                else:
                    rows_written = await self._export_xlsx(
                        str(temporary),
                        profile,
                        query,
                        output_columns,
                        progress,
                        cancelled,
                    )
            if cancelled and cancelled():
                raise asyncio.CancelledError
            os.replace(temporary, target)
        except BaseException:
            temporary.unlink(missing_ok=True)
            raise
        return ExportResult(
            collection=params.collection,
            format=fmt,
            rows_written=rows_written,
            schema_revision=profile.capability_hash,
            capability_hash=profile.capability_hash,
            output_display_name=target.name,
        )

    async def generate_template(
        self,
        collection: str,
        grant_id: str,
    ) -> TemplateResult:
        profile = self._profile(collection)
        path = self._resolve_path(
            grant_id,
            purpose="export_target",
            direction="write",
        )
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ExportError("workbook has no active worksheet", code="export_template_invalid")
        ws.title = collection[:31]
        # Header row: create-field names with required/readonly hints.
        headers: list[str] = []
        notes: list[str] = []
        create_fields = set(profile.create_fields)
        relations = {r.field: r for r in profile.relations}
        for field in profile.fields:
            if field in create_fields:
                headers.append(field)
                note = "required" if field in create_fields else "optional"
                if field in relations:
                    rel = relations[field]
                    note += f" (relation → {rel.related_collection}, match by {', '.join(rel.display_fields) or 'primary key'})"
                notes.append(note)
        if headers:
            ws.append(headers)
            ws.append(notes)
        wb.save(path)
        wb.close()
        return TemplateResult(
            collection=collection,
            grant_id=grant_id,
            display_name=Path(path).name,
        )

    # ------------------------------------------------------------------
    # Paging + streaming
    # ------------------------------------------------------------------

    async def _export_with_lookups(
        self,
        path: str,
        profile: CollectionProfile,
        params: ExportParams,
        base_columns: list[str],
        progress: Callable[[int, int, str], Awaitable[None]] | None,
        cancelled: Callable[[], bool] | None,
    ) -> int:
        assert self._lookup_provider is not None
        assert params.lookup_revision is not None
        offset = 0
        written = 0
        first = await self._lookup_provider.query_page(
            collection=profile.collection,
            fields=base_columns,
            lookup_ids=params.lookup_ids,
            lookup_revision=params.lookup_revision,
            query=params.query,
            offset=0,
            limit=EXPORT_PAGE_SIZE,
        )
        requested = set(params.lookup_ids)
        returned = {column.lookup_id for column in first.columns}
        if returned != requested:
            raise ExportError(
                "authoritative Lookup response did not describe every requested Lookup",
                code="lookup_export_columns_mismatch",
            )
        if first.lookup_revision != params.lookup_revision:
            raise ExportError(
                "Lookup definitions changed before export",
                code="lookup_revision_mismatch",
            )
        lookup_columns = [column.field_key for column in first.columns]
        if len(lookup_columns) != len(set(lookup_columns)):
            raise ExportError(
                "authoritative Lookup response contains duplicate field keys",
                code="lookup_export_columns_invalid",
            )
        columns = [*base_columns, *[col for col in lookup_columns if col not in base_columns]]
        page = first

        with ExitStack() as resources:
            if params.format == "csv":
                sink = resources.enter_context(open(path, "w", encoding="utf-8-sig", newline=""))
                writer: Any = csv.writer(sink)
                writer.writerow(columns)
                workbook = None
            else:
                from openpyxl import Workbook

                workbook = Workbook(write_only=True)
                resources.callback(workbook.close)
                writer = workbook.create_sheet(title=profile.collection[:31])
                writer.append(columns)
            while True:
                for row in page.rows:
                    rendered = [_render_cell(row, column) for column in columns]
                    if params.format == "csv":
                        writer.writerow(rendered)
                    else:
                        writer.append(rendered)
                    written += 1
                if progress:
                    await progress(written, page.filtered_rows, f"exported {written} rows")
                if len(page.rows) < EXPORT_PAGE_SIZE or (cancelled and cancelled()):
                    break
                offset += EXPORT_PAGE_SIZE
                page = await self._lookup_provider.query_page(
                    collection=profile.collection,
                    fields=base_columns,
                    lookup_ids=params.lookup_ids,
                    lookup_revision=params.lookup_revision,
                    query=params.query,
                    offset=offset,
                    limit=EXPORT_PAGE_SIZE,
                )
                if page.lookup_revision != params.lookup_revision:
                    raise ExportError(
                        "Lookup definitions changed during export",
                        code="lookup_revision_mismatch",
                    )
            if workbook is not None:
                workbook.save(path)
        return written

    async def _export_csv(
        self,
        path: str,
        profile: CollectionProfile,
        query: TableQuery,
        columns: list[str],
        progress: Callable[[int, int, str], Awaitable[None]] | None,
        cancelled: Callable[[], bool] | None,
    ) -> int:
        total_estimate = 0
        written = 0
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(columns)
            offset = 0
            while True:
                if cancelled and cancelled():
                    break
                page_query = query.model_copy(update={"offset": offset, "limit": EXPORT_PAGE_SIZE})
                page = await self._query_port.query_page(
                    table_id=profile.collection,
                    query=page_query.model_dump(mode="json", by_alias=True, exclude_none=True),
                )
                if offset == 0:
                    total_estimate = page.filtered_rows
                for row in page.rows:
                    writer.writerow([_render_cell(row, col) for col in columns])
                    written += 1
                if progress:
                    await progress(written, total_estimate, f"exported {written} rows")
                if len(page.rows) < EXPORT_PAGE_SIZE:
                    break
                offset += EXPORT_PAGE_SIZE
        return written

    async def _export_xlsx(
        self,
        path: str,
        profile: CollectionProfile,
        query: TableQuery,
        columns: list[str],
        progress: Callable[[int, int, str], Awaitable[None]] | None,
        cancelled: Callable[[], bool] | None,
    ) -> int:
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=profile.collection[:31])
        ws.append(columns)
        written = 0
        total_estimate = 0
        offset = 0
        while True:
            if cancelled and cancelled():
                break
            page_query = query.model_copy(update={"offset": offset, "limit": EXPORT_PAGE_SIZE})
            page = await self._query_port.query_page(
                table_id=profile.collection,
                query=page_query.model_dump(mode="json", by_alias=True, exclude_none=True),
            )
            if offset == 0:
                total_estimate = page.filtered_rows
            for row in page.rows:
                ws.append([_render_cell(row, col) for col in columns])
                written += 1
            if progress:
                await progress(written, total_estimate, f"exported {written} rows")
            if len(page.rows) < EXPORT_PAGE_SIZE:
                break
            offset += EXPORT_PAGE_SIZE
        wb.save(path)
        wb.close()
        return written

    def _profile(self, collection: str) -> CollectionProfile:
        profile = self._profiles.get(collection)
        if profile is None:
            raise ExportError(
                f"collection {collection!r} is not in the product schema",
                code="schema_unknown",
            )
        return profile


def _render_cell(row: dict[str, Any], column: str) -> Any:
    """Render a cell value, flattening nested relation fields (``a.b``)."""
    if "." in column:
        first, rest = column.split(".", 1)
        nested = row.get(first)
        if isinstance(nested, dict):
            return _render_cell(nested, rest)
        return nested if nested is not None else ""
    value = row.get(column)
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    if isinstance(value, (bytes, bytearray, memoryview)):
        # QueryPort exposes attachment manifests. Binary material is never
        # embedded into ordinary CSV/XLSX exports; a future ZIP exporter can
        # consume the manifest through a separate capability.
        return '{"kind":"binary_omitted"}'
    return value


def _safe_int(value: Any) -> int | None:
    return value if isinstance(value, int) else None


__all__ = [
    "EXPORT_PAGE_SIZE",
    "AuthoritativeLookupColumn",
    "AuthoritativeLookupExportPage",
    "AuthoritativeLookupExportProvider",
    "ExportError",
    "ExportService",
    "QueryPagePort",
]

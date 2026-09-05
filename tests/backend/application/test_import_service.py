"""Provider-neutral import normalization and atomic-apply tests."""

from __future__ import annotations

import csv
import json
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import httpx
import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904

from backend.adapters.pocketbase.client import PocketBaseClient
from backend.adapters.pocketbase.transport import PocketBaseConfig, StdlibPocketBaseTransport
from backend.application.import_service import (
    MAX_ATOMIC_IMPORT_ROWS,
    ImportFlowError,
    ImportService,
    RelationImportBatchResult,
    RelationImportTarget,
    SourceFile,
    auto_map_columns,
)
from backend.application.paste_service import PasteError
from backend.contracts.data_io import (
    ApplyImportParams,
    ImportColumnMapping,
    PreviewImportParams,
)
from backend.contracts.data_profile import CollectionProfile, RelationProfile
from backend.contracts.paste import ApplyPasteResult

FIELD_VALUE_CORPUS_PATH = (
    Path(__file__).resolve().parents[3]
    / "contracts"
    / "schema-v2"
    / "fixtures"
    / "field-value-entry-corpus.json"
)


def _field_value_corpus() -> list[dict[str, Any]]:
    payload = json.loads(FIELD_VALUE_CORPUS_PATH.read_text(encoding="utf-8"))
    cases = payload["cases"]
    assert cases
    return cases


class FakeProductMutationPort:
    def __init__(self, result: ApplyPasteResult | None = None) -> None:
        self.result = result or ApplyPasteResult(
            collection="vibetable_demo",
            outcome="committed",
            created_row_keys=["created-1", "created-2"],
            request_id="request-1",
        )
        self.calls: list[dict[str, Any]] = []
        self.preview_calls: list[dict[str, Any]] = []

    async def preview_import(
        self,
        *,
        collection: str,
        schema_revision: str,
        rows: list[dict[str, Any]],
        row_modes: list[str] | None = None,
    ) -> dict[str, Any]:
        self.preview_calls.append(
            {
                "collection": collection,
                "schema_revision": schema_revision,
                "rows": [dict(row) for row in rows],
                "row_modes": row_modes,
            }
        )
        normalized: list[dict[str, Any]] = []
        for row in rows:
            values = dict(row)
            diagnostics: list[dict[str, str]] = []
            if "amount" in values:
                raw_amount = values["amount"]
                if raw_amount in (None, ""):
                    values["amount"] = None
                else:
                    try:
                        values["amount"] = float(str(raw_amount).replace("$", "").replace(",", ""))
                    except ValueError:
                        values.pop("amount")
                        diagnostics.append(
                            {
                                "field": "amount",
                                "code": "field.value.invalid",
                                "message": "value is not a finite number",
                            }
                        )
            normalized.append({"values": values, "diagnostics": diagnostics})
        return {"contract": "vibetable.import-preview.v1", "rows": normalized}

    async def apply(self, **kwargs: Any) -> ApplyPasteResult:
        self.calls.append(kwargs)
        return self.result


class FakeRelationProvider:
    def __init__(self, matches: dict[str, list[Any]]) -> None:
        self.matches = matches
        self.inspected: list[tuple[str, str]] = []
        self.applied: list[dict[str, Any]] = []

    async def inspect_mapping(
        self,
        *,
        collection: str,
        target_field: str,
        relation_id: str,
        match_field: str,
    ) -> RelationImportTarget:
        del collection
        self.inspected.append((relation_id, match_field))
        if match_field == "title":
            raise ValueError("match field is not unique")
        return RelationImportTarget(
            relation_id=relation_id,
            target_field=target_field,
            target_collection="contracts",
            target_primary_key="id",
            match_field=match_field,
        )

    async def find_exact(
        self,
        target: RelationImportTarget,
        value: Any,
    ) -> list[Any]:
        del target
        return self.matches.get(str(value), [])

    async def apply_chunk(self, **kwargs: Any) -> RelationImportBatchResult:
        self.applied.append(kwargs)
        return RelationImportBatchResult(
            created_row_keys=["source-1"],
            updated_row_keys=[],
            request_id="relation-request-1",
        )


def _profile(*, relation: bool = False) -> CollectionProfile:
    relations = (
        [
            RelationProfile(
                relation_id="rel_contract",
                field="contract",
                kind="m2o",
                related_collection="contracts",
                display_fields=["number"],
            )
        ]
        if relation
        else []
    )
    relation_fields = ["contract"] if relation else []
    return CollectionProfile(
        collection="vibetable_demo",
        schema_revision="schema-1",
        fields=[
            "id",
            "status",
            "number",
            "title",
            "amount",
            "signed_on",
            "date_updated",
            *relation_fields,
        ],
        field_schemas={
            "number": {"dataType": "shortText", "constraints": []},
            "title": {"dataType": "shortText", "constraints": []},
            "amount": {"dataType": "float", "constraints": []},
            "signed_on": {"dataType": "date", "constraints": []},
            "status": {
                "dataType": "select",
                "constraints": [
                    {
                        "kind": "enum",
                        "options": [{"value": "active"}, {"value": "archived"}],
                    }
                ],
            },
        },
        create_fields=[
            "id",
            "status",
            "number",
            "title",
            "amount",
            "signed_on",
            *relation_fields,
        ],
        update_fields=[
            "status",
            "number",
            "title",
            "amount",
            "signed_on",
            *relation_fields,
        ],
        relations=relations,
    )


def _write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(header)
        writer.writerows(rows)


def _service(
    path: Path,
    *,
    profile: CollectionProfile | None = None,
    mutation: FakeProductMutationPort | None = None,
    relation_provider: FakeRelationProvider | None = None,
    consumed: list[str] | None = None,
    clock: Any = None,
) -> tuple[ImportService, FakeProductMutationPort]:
    profile = profile or _profile()
    mutation = mutation or FakeProductMutationPort()
    kwargs: dict[str, Any] = {}
    if clock is not None:
        kwargs["clock"] = clock
    return (
        ImportService(
            client=object(),
            auth=object(),
            bulk=mutation,
            profiles={profile.collection: profile},
            resolve_path=lambda _grant, **_kwargs: str(path),
            consume_grant=lambda grant: consumed.append(grant) if consumed is not None else None,
            relation_provider=relation_provider,
            **kwargs,
        ),
        mutation,
    )


def test_import_error_exposes_only_structured_product_data() -> None:
    error = ImportFlowError(
        "internal detail",
        code="schema_mismatch",
        data={"currentSchemaRevision": "schema-2"},
    )
    assert error.rpc_error_data == {
        "code": "schema_mismatch",
        "currentSchemaRevision": "schema-2",
    }


def test_source_file_rejects_csv_beyond_atomic_limit_and_other_formats(tmp_path: Path) -> None:
    csv_path = tmp_path / "source.csv"
    _write_csv(csv_path, ["number", "title"], [["1", "one"], ["2", "two"]])

    with pytest.raises(ImportFlowError) as overflow:
        SourceFile(str(csv_path)).read_header_and_rows(max_rows=1)
    assert overflow.value.code == "import_row_limit"
    assert overflow.value.rpc_error_data == {
        "code": "import_row_limit",
        "maxRows": 1,
    }

    unsupported = tmp_path / "source.txt"
    unsupported.write_text("x", encoding="utf-8")
    with pytest.raises(ImportFlowError) as error:
        SourceFile(str(unsupported)).read_header_and_rows()
    assert error.value.code == "import_unsupported_format"


@pytest.mark.asyncio
async def test_xlsx_native_dates_reach_import_http_as_json_scalars(tmp_path: Path) -> None:
    path = tmp_path / "native-dates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["signed_on", "recorded_at", "title", "amount", "enabled", "blank"])
    sheet.append([date(2026, 8, 29), datetime(2026, 8, 29, 14, 5, 6, 123000), "日期", 0, False])
    workbook.save(path)
    workbook.close()
    header, rows, _ = SourceFile(str(path)).read_header_and_rows()
    received: list[dict[str, Any]] = []

    def receive(request: httpx.Request) -> httpx.Response:
        received.append(json.loads(request.content))
        return httpx.Response(200, json={"contract": "vibetable.import-preview.v1", "rows": []})

    transport = StdlibPocketBaseTransport(
        PocketBaseConfig(base_url="http://127.0.0.1:1", session_secret="0" * 64),
        http_transport=httpx.MockTransport(receive),
    )
    client = PocketBaseClient(transport=transport, session_secret="0" * 64)
    await client.preview_import(
        {
            "contract": "vibetable.import-preview.v1",
            "tableId": "native-dates",
            "schemaRevision": "1",
            "rows": [{"values": dict(zip(header, rows[0], strict=True)), "mode": "insert"}],
        }
    )
    assert received[0]["rows"][0]["values"] == {
        "signed_on": "2026-08-29",
        "recorded_at": "2026-08-29 14:05:06.123000",
        "title": "日期",
        "amount": 0,
        "enabled": False,
        "blank": None,
    }


@pytest.mark.parametrize(
    ("value", "number_format", "code"),
    [
        (59, "yyyy-mm-dd", "import_ambiguous_excel_date"),
        (60, "yyyy-mm-dd", "import_ambiguous_excel_date"),
        (time(12, 30), "hh:mm:ss", "import_unsupported_excel_time"),
        (timedelta(days=1, hours=2), "[h]:mm:ss", "import_unsupported_excel_time"),
    ],
)
def test_xlsx_unrepresentable_native_dates_have_explicit_errors(
    tmp_path: Path, value: object, number_format: str, code: str
) -> None:
    path = tmp_path / "unsupported-native-date.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["value"])
    sheet.append([value])
    sheet["A2"].number_format = number_format
    workbook.save(path)
    workbook.close()

    with pytest.raises(ImportFlowError) as error:
        SourceFile(str(path)).read_header_and_rows()
    assert error.value.code == code
    assert error.value.data == {"sheet": "Sheet", "row": 2, "column": 1}
    assert "ISO" in str(error.value)


def test_xlsx_dates_preserve_hidden_time_and_formula_cache_semantics(tmp_path: Path) -> None:
    path = tmp_path / "date-formats.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["value"])
    sheet.append([datetime(2026, 8, 29)])
    sheet["A2"].number_format = "yyyy-mm-dd hh:mm:ss"
    sheet.append([datetime(2026, 8, 29, 12, 30)])
    sheet["A3"].number_format = "yyyy-mm-dd"
    sheet.append([date(2026, 8, 29)])
    sheet["A4"].number_format = "YYYY-MM-DD"
    sheet.append(["1900-02-28"])
    sheet.append(["2026-08-29T00:00:00+08:00"])
    sheet.append(["=1+1"])
    sheet.append(["=1+1"])
    sheet["A8"].data_type = "s"
    workbook.save(path)
    workbook.close()
    _, rows, _ = SourceFile(str(path)).read_header_and_rows()
    assert rows == [
        ["2026-08-29 00:00:00"],
        ["2026-08-29 12:30:00"],
        ["2026-08-29"],
        ["1900-02-28"],
        ["2026-08-29T00:00:00+08:00"],
        [None],
        ["=1+1"],
    ]


def test_xlsx_1904_epoch_does_not_reject_unambiguous_native_date(tmp_path: Path) -> None:
    path = tmp_path / "1904.xlsx"
    workbook = Workbook()
    workbook.epoch = CALENDAR_MAC_1904
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["value"])
    sheet.append([date(1900, 2, 28)])
    workbook.save(path)
    workbook.close()
    _, rows, _ = SourceFile(str(path)).read_header_and_rows()
    assert rows == [["1900-02-28"]]


def test_source_file_reads_named_xlsx_sheet_and_empty_workbook(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "First"
    sheet.append(["ignored"])
    chosen = workbook.create_sheet("Chosen")
    chosen.append(["number", "title"])
    chosen.append(["1", "one"])
    path = tmp_path / "source.xlsx"
    workbook.save(path)
    workbook.close()

    header, rows, _ = SourceFile(str(path)).read_header_and_rows(
        max_rows=1,
        sheet="Chosen",
    )
    assert header == ["number", "title"]
    assert rows == [["1", "one"]]

    empty = openpyxl.Workbook()
    empty_sheet = empty.active
    assert empty_sheet is not None
    empty_sheet.delete_rows(1, empty_sheet.max_row)
    empty_path = tmp_path / "empty.xlsx"
    empty.save(empty_path)
    empty.close()
    header, rows, _ = SourceFile(str(empty_path)).read_header_and_rows()
    assert header == []
    assert rows == []


def test_source_file_rejects_xlsx_beyond_atomic_limit(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["number"])
    sheet.append(["A-1"])
    sheet.append(["A-2"])
    path = tmp_path / "source.xlsx"
    workbook.save(path)
    workbook.close()

    with pytest.raises(ImportFlowError) as overflow:
        SourceFile(str(path)).read_header_and_rows(max_rows=1)
    assert overflow.value.code == "import_row_limit"
    assert overflow.value.rpc_error_data["maxRows"] == 1


def test_atomic_import_row_limit_matches_product_mutation_kernel() -> None:
    assert MAX_ATOMIC_IMPORT_ROWS == 1_000


def test_auto_mapping_is_case_insensitive_and_explicit_mapping_wins() -> None:
    profile = _profile()
    mapping, unmatched = auto_map_columns(
        ["Number", "Signed On", "External title"],
        profile,
        [
            ImportColumnMapping(
                source_column="External title",
                target_field="title",
            )
        ],
    )
    assert mapping == {0: "number", 1: "signed_on", 2: "title"}
    assert unmatched == []


@pytest.mark.asyncio
async def test_preview_binds_source_and_normalizes_product_fields(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    _write_csv(
        path,
        ["number", "amount", "signed_on", "unused"],
        [["A-1", "$1,234.50", "2026-07-14", "ignored"]],
    )
    service, mutation = _service(path)

    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )

    assert len(plan.source_hash) == 64
    assert plan.summary.valid_rows == 1
    assert plan.rows[0].values == {
        "number": "A-1",
        "amount": 1234.5,
        "signed_on": "2026-07-14",
    }
    assert plan.unmatched_columns == ["unused"]
    assert plan.token.token
    assert mutation.preview_calls[0]["rows"] == [
        {
            "number": "A-1",
            "amount": "$1,234.50",
            "signed_on": "2026-07-14",
        }
    ]


@pytest.mark.asyncio
async def test_import_forwards_shared_corpus_raw_values_unchanged(tmp_path: Path) -> None:
    cases = _field_value_corpus()
    path = tmp_path / "field-value-corpus.csv"
    _write_csv(
        path,
        [case["field"] for case in cases],
        [[case["rawValue"] for case in cases]],
    )
    profile = _profile()
    profile.fields.extend(case["field"] for case in cases if case["field"] not in profile.fields)
    profile.create_fields.extend(
        case["field"] for case in cases if case["field"] not in profile.create_fields
    )
    service, mutation = _service(path, profile=profile)

    await service.preview(
        PreviewImportParams(
            grant_id="grant-corpus",
            collection=profile.collection,
            schema_revision=profile.schema_revision,
        )
    )

    assert mutation.preview_calls[0]["rows"] == [
        {case["field"]: case["rawValue"] for case in cases}
    ]


@pytest.mark.asyncio
async def test_preview_preserves_explicit_blank_cells_as_supplied(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    _write_csv(path, ["number", "amount"], [["", ""]])
    service, _ = _service(path)

    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )

    assert plan.rows[0].values == {"number": "", "amount": None}


@pytest.mark.asyncio
async def test_apply_is_one_atomic_product_mutation_and_consumes_grant(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    _write_csv(path, ["number"], [["A-1"], ["A-2"]])
    consumed: list[str] = []
    service, mutation = _service(path, consumed=consumed)
    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )

    result = await service.apply(
        ApplyImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            token=plan.token.token,
            idempotency_prefix="import-1",
        )
    )

    assert result.created_count == 2
    assert result.failed_rows == []
    assert len(result.chunks) == 1
    assert result.chunks[0].idempotency_key == "import-1-0"
    assert len(mutation.calls) == 1
    assert mutation.calls[0]["schema_revision"] == "schema-1"
    assert len(mutation.calls[0]["rows"]) == 2
    assert consumed == ["grant-1"]

    with pytest.raises(ImportFlowError) as replay:
        await service.apply(
            ApplyImportParams(
                grant_id="grant-1",
                collection="vibetable_demo",
                token=plan.token.token,
            )
        )
    assert replay.value.code == "import_token_consumed"


@pytest.mark.asyncio
async def test_thousand_row_import_uses_one_atomic_product_mutation(tmp_path: Path) -> None:
    path = tmp_path / "thousand.csv"
    _write_csv(
        path,
        ["number"],
        [[f"A-{index:04d}"] for index in range(1_000)],
    )
    mutation = FakeProductMutationPort(
        ApplyPasteResult(
            collection="vibetable_demo",
            outcome="committed",
            created_row_keys=[f"row-{index:04d}" for index in range(1_000)],
            request_id="request-1000",
        )
    )
    service, _ = _service(path, mutation=mutation)
    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1000",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )

    result = await service.apply(
        ApplyImportParams(
            grant_id="grant-1000",
            collection="vibetable_demo",
            token=plan.token.token,
            idempotency_prefix="import-1000",
        )
    )

    assert result.created_count == 1_000
    assert len(mutation.calls) == 1
    assert len(mutation.calls[0]["rows"]) == 1_000


@pytest.mark.asyncio
async def test_failed_atomic_apply_commits_nothing_and_keeps_grant(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    _write_csv(path, ["number"], [["A-1"], ["A-2"]])
    consumed: list[str] = []
    mutation = FakeProductMutationPort(
        ApplyPasteResult(
            collection="vibetable_demo",
            outcome="conflict",
        )
    )
    service, _ = _service(path, consumed=consumed, mutation=mutation)
    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )

    progress_messages: list[str] = []

    async def progress(_done: int, _total: int, message: str) -> None:
        progress_messages.append(message)

    result = await service.apply(
        ApplyImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            token=plan.token.token,
        ),
        progress=progress,
    )

    assert result.created_count == result.updated_count == 0
    assert result.failed_rows == [2, 3]
    assert consumed == []
    assert progress_messages == ["atomic import failed [import_conflict]"]


@pytest.mark.asyncio
async def test_failed_atomic_apply_surfaces_safe_product_path_and_message(
    tmp_path: Path,
) -> None:
    class ProductCauseError(Exception):
        path = "payload"

    class FailingMutation(FakeProductMutationPort):
        async def apply(self, **kwargs: Any) -> ApplyPasteResult:
            del kwargs
            try:
                raise ProductCauseError("Invalid JSON value.")
            except ProductCauseError as cause:
                raise PasteError(
                    "PocketBase rejected the mutation",
                    code="mutation.validation.failed",
                ) from cause

    path = tmp_path / "source.csv"
    _write_csv(path, ["number"], [["A-1"]])
    service, _ = _service(path, mutation=FailingMutation())
    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )
    progress_messages: list[str] = []

    async def progress(_done: int, _total: int, message: str) -> None:
        progress_messages.append(message)

    result = await service.apply(
        ApplyImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            token=plan.token.token,
        ),
        progress=progress,
    )

    assert result.failed_rows == [2]
    assert progress_messages == [
        "atomic import failed [mutation.validation.failed]: at payload: Invalid JSON value."
    ]


@pytest.mark.asyncio
async def test_relation_preview_requires_stable_identity_and_exact_unique_match(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.csv"
    _write_csv(path, ["contract"], [["C-1"], ["missing"], ["duplicate"]])
    provider = FakeRelationProvider(
        {
            "C-1": ["contract-1"],
            "duplicate": ["contract-2", "contract-3"],
        }
    )
    service, _ = _service(
        path,
        profile=_profile(relation=True),
        relation_provider=provider,
    )

    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
            column_mapping=[
                ImportColumnMapping(
                    source_column="contract",
                    target_field="contract",
                    relation_id="rel_contract",
                    match_field="number",
                )
            ],
        )
    )

    assert plan.rows[0].values["contract"] == "contract-1"
    assert plan.rows[0].relation_resolutions[0].state == "matched"
    assert plan.rows[1].diagnostics[0].code == "relation_match_not_found"
    assert plan.rows[2].diagnostics[0].code == "relation_match_ambiguous"
    assert provider.inspected == [("rel_contract", "number")]


@pytest.mark.asyncio
async def test_expired_import_token_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    _write_csv(path, ["number"], [["A-1"]])
    now = [100.0]
    service, _ = _service(path, clock=lambda: now[0])
    plan = await service.preview(
        PreviewImportParams(
            grant_id="grant-1",
            collection="vibetable_demo",
            schema_revision="schema-1",
        )
    )
    now[0] += 601

    with pytest.raises(ImportFlowError) as error:
        await service.apply(
            ApplyImportParams(
                grant_id="grant-1",
                collection="vibetable_demo",
                token=plan.token.token,
            )
        )
    assert error.value.code == "import_token_expired"

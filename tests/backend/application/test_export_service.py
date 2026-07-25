"""C1 export service tests.

Covers CSV/XLSX streaming export with paging, relation column flattening,
cooperative cancellation, and template generation.
"""

from __future__ import annotations

import asyncio
import csv
import json
from dataclasses import dataclass
from typing import Any

import pytest

from backend.application.export_service import (
    AuthoritativeLookupColumn,
    AuthoritativeLookupExportPage,
    ExportError,
    ExportService,
)
from backend.contracts.data_io import ExportParams
from backend.contracts.data_profile import CollectionProfile


@dataclass
class FakePage:
    rows: list[dict[str, Any]]
    filtered_rows: int
    total_rows: int


class FakeQueryPort:
    def __init__(
        self, pages: list[list[dict[str, Any]]], meta: dict[str, Any] | None = None
    ) -> None:
        self._pages = list(pages)
        self._meta = meta or {"filter_count": sum(len(p) for p in pages)}
        self.calls: list[dict[str, Any]] = []

    async def query_page(self, *, table_id: str, query: dict[str, Any]) -> FakePage:
        self.calls.append({"table_id": table_id, "query": query})
        rows = self._pages.pop(0) if self._pages else []
        total = self._meta.get("filter_count", self._meta.get("total_count", len(rows)))
        return FakePage(rows=rows, filtered_rows=total, total_rows=total)


def _manifest() -> dict[str, CollectionProfile]:
    profile = CollectionProfile.model_validate(
        {
            "collection": "vibetable_demo",
            "primary_key": "id",
            "fields": [
                "id",
                "status",
                "number",
                "title",
                "amount",
                "owner",
                "payload",
                "attachments",
                "date_updated",
            ],
            "create_fields": ["number", "title", "amount", "owner"],
            "update_fields": ["number", "title", "amount"],
            "archive_field": "status",
            "archive_value": "archived",
            "restore_value": "active",
            "date_updated_field": "date_updated",
            "relations": [
                {
                    "field": "owner",
                    "kind": "m2o",
                    "related_collection": "users",
                    "display_fields": ["first_name", "last_name"],
                }
            ],
        }
    )
    return {profile.collection: profile}


def _service(
    query_port: FakeQueryPort,
    profiles: dict[str, CollectionProfile],
    path: str,
) -> ExportService:
    return ExportService(
        query_port=query_port,
        profiles=profiles,
        resolve_path=lambda _g, *, purpose, direction: path,
    )


class FakeLookupExportProvider:
    def __init__(self, pages: list[list[dict[str, Any]]], *, revision: str = "lookup-r1") -> None:
        self.pages = list(pages)
        self.revision = revision
        self.calls: list[dict[str, Any]] = []

    async def query_page(self, **kwargs: Any) -> AuthoritativeLookupExportPage:
        self.calls.append(kwargs)
        rows = self.pages.pop(0) if self.pages else []
        return AuthoritativeLookupExportPage(
            rows=rows,
            columns=[AuthoritativeLookupColumn("contract_price", "contract_price")],
            filtered_rows=sum(len(page) for page in self.pages) + len(rows),
            lookup_revision=self.revision,
        )


def _lookup_service(
    query_port: FakeQueryPort,
    profiles: dict[str, CollectionProfile],
    path: str,
    provider: Any,
) -> ExportService:
    return ExportService(
        query_port=query_port,
        profiles=profiles,
        resolve_path=lambda _g, *, purpose, direction: path,
        lookup_provider=provider,
    )


@pytest.mark.asyncio
async def test_export_csv_streams_all_pages(tmp_path: Any) -> None:
    from backend.application.export_service import EXPORT_PAGE_SIZE

    manifest = _manifest()
    # Page 1 is full (EXPORT_PAGE_SIZE rows); page 2 is partial (triggers stop).
    page1 = [{"id": str(i), "number": f"A-{i}"} for i in range(EXPORT_PAGE_SIZE)]
    page2 = [{"id": "998", "number": "A-998"}]
    transport = FakeQueryPort([page1, page2])
    path = tmp_path / "out.csv"
    service = _service(transport, manifest, str(path))
    result = await service.export(
        ExportParams(grant_id="g1", collection="vibetable_demo", query={}, format="csv")
    )
    assert result.rows_written == EXPORT_PAGE_SIZE + 1
    with open(str(path), encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        assert "id" in header
        assert "number" in header
        rows = list(reader)
    assert len(rows) == EXPORT_PAGE_SIZE + 1


@pytest.mark.asyncio
async def test_export_csv_with_relations_adds_display_columns(tmp_path: Any) -> None:
    manifest = _manifest()
    page1 = [{"id": "1", "number": "A-1", "owner": {"first_name": "Ada", "last_name": "Lovelace"}}]
    transport = FakeQueryPort([page1])
    path = tmp_path / "out.csv"
    service = _service(transport, manifest, str(path))
    result = await service.export(
        ExportParams(
            grant_id="g1",
            collection="vibetable_demo",
            query={},
            format="csv",
            include_relations=True,
        )
    )
    assert result.rows_written == 1
    with open(str(path), encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        assert "owner.first_name" in header
        row = next(reader)
        idx = header.index("owner.first_name")
        assert row[idx] == "Ada"


@pytest.mark.asyncio
async def test_export_xlsx_writes_rows(tmp_path: Any) -> None:
    manifest = _manifest()
    page1 = [{"id": "1", "number": "A-1"}]
    transport = FakeQueryPort([page1])
    path = tmp_path / "out.xlsx"
    service = _service(transport, manifest, str(path))
    result = await service.export(
        ExportParams(grant_id="g1", collection="vibetable_demo", query={}, format="xlsx")
    )
    assert result.rows_written == 1
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True)
    ws = wb.active
    rows = list(ws.values)
    wb.close()
    assert len(rows) == 2  # header + 1 data


@pytest.mark.asyncio
async def test_export_cancellation_stops_at_page_boundary(tmp_path: Any) -> None:
    manifest = _manifest()
    page1 = [{"id": "1", "number": "A-1"}, {"id": "2", "number": "A-2"}]
    page2 = [{"id": "3", "number": "A-3"}, {"id": "4", "number": "A-4"}]
    transport = FakeQueryPort([page1, page2])
    path = tmp_path / "out.csv"
    path.write_text("previous complete export", encoding="utf-8")
    service = _service(transport, manifest, str(path))
    # Cancel after the first page.
    call_count = [0]

    def is_cancelled() -> bool:
        call_count[0] += 1
        return call_count[0] > 1

    with pytest.raises(asyncio.CancelledError):
        await service.export(
            ExportParams(grant_id="g1", collection="vibetable_demo", query={}, format="csv"),
            cancelled=is_cancelled,
        )
    assert path.read_text(encoding="utf-8") == "previous complete export"
    assert not list(tmp_path.glob("*.tmp"))


@pytest.mark.asyncio
async def test_generate_template_writes_headers_and_notes(tmp_path: Any) -> None:
    manifest = _manifest()
    transport = FakeQueryPort([])
    path = tmp_path / "template.xlsx"
    service = _service(transport, manifest, str(path))
    result = await service.generate_template("vibetable_demo", "g1")
    assert result.display_name == "template.xlsx"
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True)
    ws = wb.active
    rows = list(ws.values)
    wb.close()
    header = rows[0]
    assert "number" in header
    assert "title" in header


@pytest.mark.asyncio
async def test_lookup_export_requires_authoritative_provider_without_page_fallback(
    tmp_path: Any,
) -> None:
    manifest = _manifest()
    transport = FakeQueryPort([[{"id": "current-page", "contract_price": 1}]])
    path = tmp_path / "lookups.csv"
    service = _service(transport, manifest, str(path))

    with pytest.raises(ExportError) as caught:
        await service.export(
            ExportParams(
                grant_id="g1",
                collection="vibetable_demo",
                lookup_ids=["contract_price"],
                lookup_revision="lookup-r1",
            )
        )

    assert caught.value.code == "lookup_export_provider_missing"
    assert transport.calls == []
    assert not path.exists()


@pytest.mark.asyncio
async def test_lookup_export_streams_authoritative_full_dataset_and_revision(tmp_path: Any) -> None:
    from backend.application.export_service import EXPORT_PAGE_SIZE

    manifest = _manifest()
    first = [
        {"id": str(index), "number": f"A-{index}", "contract_price": index * 10}
        for index in range(EXPORT_PAGE_SIZE)
    ]
    second = [{"id": "last", "number": "A-last", "contract_price": 999}]
    provider = FakeLookupExportProvider([first, second])
    transport = FakeQueryPort([])
    path = tmp_path / "lookups.csv"
    service = _lookup_service(transport, manifest, str(path), provider)

    result = await service.export(
        ExportParams(
            grant_id="g1",
            collection="vibetable_demo",
            query={"filters": [], "sorts": [{"field": "contract_price", "direction": "desc"}]},
            lookup_ids=["contract_price"],
            lookup_revision="lookup-r1",
        )
    )

    assert result.rows_written == EXPORT_PAGE_SIZE + 1
    assert [call["offset"] for call in provider.calls] == [0, EXPORT_PAGE_SIZE]
    assert all(call["lookup_revision"] == "lookup-r1" for call in provider.calls)
    assert provider.calls[0]["query"]["sorts"][0]["field"] == "contract_price"
    assert transport.calls == []
    with open(path, encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        rows = list(reader)
    assert "contract_price" in header
    assert len(rows) == EXPORT_PAGE_SIZE + 1
    assert rows[-1][header.index("contract_price")] == "999"


@pytest.mark.asyncio
async def test_lookup_export_rejects_revision_drift(tmp_path: Any) -> None:
    manifest = _manifest()
    provider = FakeLookupExportProvider([], revision="lookup-r2")
    path = tmp_path / "lookups.csv"
    service = _lookup_service(FakeQueryPort([]), manifest, str(path), provider)

    with pytest.raises(ExportError) as caught:
        await service.export(
            ExportParams(
                grant_id="g1",
                collection="vibetable_demo",
                lookup_ids=["contract_price"],
                lookup_revision="lookup-r1",
            )
        )

    assert caught.value.code == "lookup_revision_mismatch"
    assert not path.exists()


@pytest.mark.asyncio
async def test_export_renders_json_and_attachment_manifest_without_binary_data(
    tmp_path: Any,
) -> None:
    manifest = _manifest()
    page = [
        {
            "id": "1",
            "payload": {"nested": [1, True, None], "label": "原样"},
            "attachments": [
                {
                    "storedName": "report_abc.pdf",
                    "originalName": "报告.pdf",
                    "size": 42,
                    "contentType": "application/pdf",
                }
            ],
        }
    ]
    path = tmp_path / "manifest.csv"
    service = _service(FakeQueryPort([page]), manifest, str(path))

    await service.export(
        ExportParams(grant_id="g1", collection="vibetable_demo", query={}, format="csv")
    )

    with open(path, encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        row = next(reader)
    assert json.loads(row["payload"]) == page[0]["payload"]
    assert json.loads(row["attachments"]) == page[0]["attachments"]
    assert "bytes" not in row["attachments"].lower()

"""CSV/XLSX interoperability contracts exercised through the public export seam."""

from __future__ import annotations

import csv
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Literal

import pytest
from openpyxl import load_workbook
from pydantic import JsonValue

from backend.application.export_service import (
    AuthoritativeLookupColumn,
    AuthoritativeLookupExportPage,
    ExportService,
)
from backend.contracts.data_io import ExportParams
from backend.contracts.data_profile import CollectionProfile


class _OnePageQueryPort:
    def __init__(self, row: Mapping[str, JsonValue]) -> None:
        self.rows = [dict(row)]
        self.filtered_rows = 1
        self.total_rows = 1

    async def query_page(self, *, table_id: str, query: dict[str, object]) -> _OnePageQueryPort:
        del table_id, query
        return self


class _OnePageLookupProvider:
    def __init__(self, row: Mapping[str, JsonValue]) -> None:
        self._row = dict(row)

    async def query_page(
        self,
        *,
        collection: str,
        fields: list[str],
        lookup_ids: list[str],
        lookup_revision: str,
        query: dict[str, object],
        offset: int,
        limit: int,
    ) -> AuthoritativeLookupExportPage:
        del collection, fields, lookup_ids, lookup_revision, query, offset, limit
        return AuthoritativeLookupExportPage(
            rows=[self._row],
            columns=[AuthoritativeLookupColumn("lookup-formula", "lookup_formula")],
            filtered_rows=1,
            lookup_revision="lookup-r1",
        )


def _profile(fields: list[str]) -> CollectionProfile:
    return CollectionProfile.model_validate(
        {
            "collection": "interoperability",
            "primary_key": "id",
            "fields": fields,
            "create_fields": [],
            "update_fields": [],
            "archive_field": None,
            "date_updated_field": None,
        }
    )


async def _export(
    row: Mapping[str, JsonValue],
    target: Path,
    export_format: Literal["csv", "xlsx"],
) -> None:
    service = ExportService(
        query_port=_OnePageQueryPort(row),
        profiles={"interoperability": _profile(list(row))},
        resolve_path=lambda _grant, *, purpose, direction: str(target),
    )
    await service.export(
        ExportParams(
            grant_id="export-target",
            collection="interoperability",
            query={},
            format=export_format,
        )
    )


@pytest.mark.asyncio
async def test_csv_and_xlsx_preserve_interoperable_multi_field_semantics(tmp_path: Path) -> None:
    row: dict[str, JsonValue] = {
        "id": "record-1",
        "blank": None,
        "empty_text": "",
        "zero": 0,
        "disabled": False,
        "title": '東京, München — "原样"',
        "occurred_at": "2026-08-28T12:34:56+05:30",
        "status": "approved",
        "owner": {"id": "user-1", "name": "Ada"},
        "payload": {"enabled": False, "items": [0, "雪"]},
        "formula_text": "=1+1",
    }
    csv_path = tmp_path / "interoperability.csv"
    xlsx_path = tmp_path / "interoperability.xlsx"

    await _export(row, csv_path, "csv")
    await _export(row, xlsx_path, "xlsx")

    with csv_path.open(encoding="utf-8-sig", newline="") as stream:
        csv_row = next(csv.DictReader(stream))
    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    worksheet = workbook.active
    assert worksheet is not None
    worksheet_rows = worksheet.iter_rows()
    headers = [cell.value for cell in next(worksheet_rows)]
    cells = list(next(worksheet_rows))
    xlsx_row = dict(zip(headers, (cell.value for cell in cells), strict=True))
    formula_cell = cells[headers.index("formula_text")]
    workbook.close()

    expected = {
        "id": "record-1",
        "blank": "",
        "empty_text": "",
        "zero": "0",
        "disabled": "false",
        "title": '東京, München — "原样"',
        "occurred_at": "2026-08-28T12:34:56+05:30",
        "status": "approved",
        "owner": '{"id":"user-1","name":"Ada"}',
        "payload": '{"enabled":false,"items":[0,"雪"]}',
        "formula_text": "=1+1",
    }

    def normalize(exported: Iterable[tuple[object, object]]) -> dict[str, str]:
        normalized = {str(key): "" if value is None else str(value) for key, value in exported}
        normalized["disabled"] = normalized["disabled"].lower()
        return normalized

    assert normalize(csv_row.items()) == expected
    assert normalize(xlsx_row.items()) == expected
    assert formula_cell.value == "=1+1"
    assert formula_cell.data_type == "s"


@pytest.mark.asyncio
async def test_xlsx_lookup_export_preserves_base_and_lookup_formula_text(tmp_path: Path) -> None:
    target = tmp_path / "lookup-interoperability.xlsx"
    row: dict[str, JsonValue] = {
        "id": "record-1",
        "formula_text": "=BASE_VALUE",
        "lookup_formula": "=LOOKUP_VALUE",
    }
    service = ExportService(
        query_port=_OnePageQueryPort(row),
        profiles={"interoperability": _profile(["id", "formula_text"])},
        resolve_path=lambda _grant, *, purpose, direction: str(target),
        lookup_provider=_OnePageLookupProvider(row),
    )

    await service.export(
        ExportParams(
            grant_id="export-target",
            collection="interoperability",
            query={},
            format="xlsx",
            lookup_ids=["lookup-formula"],
            lookup_revision="lookup-r1",
        )
    )

    workbook = load_workbook(target, read_only=True, data_only=False)
    worksheet = workbook.active
    assert worksheet is not None
    worksheet_rows = worksheet.iter_rows()
    headers = [cell.value for cell in next(worksheet_rows)]
    cells = list(next(worksheet_rows))
    by_header = dict(zip(headers, cells, strict=True))
    workbook.close()

    assert by_header["formula_text"].value == "=BASE_VALUE"
    assert by_header["formula_text"].data_type == "s"
    assert by_header["lookup_formula"].value == "=LOOKUP_VALUE"
    assert by_header["lookup_formula"].data_type == "s"

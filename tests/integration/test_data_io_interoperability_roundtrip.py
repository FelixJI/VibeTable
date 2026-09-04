from __future__ import annotations

import csv
import json
import os
import subprocess
import tempfile
from collections.abc import Iterator, Mapping
from pathlib import Path
from typing import NotRequired, TypedDict, cast

import pytest
from openpyxl import load_workbook

from backend.adapters.pocketbase.client import PocketBaseClient
from backend.adapters.pocketbase.data_io import ProductDataIoRuntime
from backend.adapters.pocketbase.transport import PocketBaseConfig, StdlibPocketBaseTransport
from backend.application.task_service import build_task_service
from backend.contracts.data_io import ApplyImportParams, ExportParams, PreviewImportParams
from backend.contracts.task import HostExportTargetParams, HostImportSourceParams
from scripts.build_next import RepoPaths, build_sidecar_command
from tests.integration.packaged_sidecar_matrix import (
    CLAIM_ID,
    FENCE_EPOCH,
    SESSION_EPOCH,
    WORKSPACE_ID,
    Sidecar,
    _create_field,
    _create_table,
    _create_v2_workspace,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
CORPUS_PATH = REPO_ROOT / "tests" / "fixtures" / "data-io" / "a5-falsy-container-corpus.json"


class SelectOption(TypedDict):
    optionId: str
    label: str
    color: str
    order: int
    state: str


class CorpusCase(TypedDict):
    key: str
    displayName: str
    logicalType: str
    rawCell: str
    productValue: object
    exportText: str
    selectOption: NotRequired[SelectOption]


def _load_cases() -> list[CorpusCase]:
    payload = json.loads(CORPUS_PATH.read_text(encoding="utf-8"))
    cases = cast(list[CorpusCase], payload["cases"])
    assert cases
    assert len({case["key"] for case in cases}) == len(cases)
    return cases


def _build_source_sidecar(output_dir: Path) -> Path:
    output = output_dir / ("vibetable-pb.exe" if os.name == "nt" else "vibetable-pb")
    paths = RepoPaths.default(REPO_ROOT)
    subprocess.run(
        build_sidecar_command(
            paths,
            output=output,
            commit="a5-falsy-container-roundtrip",
            build_time="2026-09-04T00:00:00Z",
        ),
        cwd=paths.sidecar_source_dir,
        check=True,
    )
    return output


@pytest.fixture(scope="module")
def source_sidecar_binary() -> Iterator[Path]:
    build_root = REPO_ROOT / "build" / "qa"
    build_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="a5-falsy-container-", dir=build_root) as run_root:
        yield _build_source_sidecar(Path(run_root))


def _field_draft(sidecar: Sidecar, table_id: str, case: CorpusCase) -> dict[str, object]:
    described = sidecar.request("GET", f"/api/vibetable/v2/field-settings/{table_id}").json()
    capability = next(
        item for item in described["capabilities"] if item["logicalType"] == case["logicalType"]
    )
    recommended = capability["recommended"]
    draft = {
        "displayName": case["displayName"],
        "help": "",
        "logicalType": case["logicalType"],
        "value": recommended["value"],
        "constraints": recommended["constraints"],
        "storage": recommended["storage"],
        "display": recommended["display"],
    }
    if recommended.get("json") is not None:
        draft["json"] = recommended["json"]
    option = case.get("selectOption")
    if option is not None:
        draft["select"] = {"options": [option]}
    return draft


def _physical_name(definition: Mapping[str, object]) -> str:
    identity = definition["identity"]
    assert isinstance(identity, dict)
    physical_name = identity["physicalName"]
    assert isinstance(physical_name, str)
    return physical_name


def _export_text(value: object, logical_type: str) -> str:
    if value is None:
        return ""
    rendered = str(value)
    return rendered.lower() if logical_type == "bool" else rendered


def _assert_strict_json(actual: object, expected: object) -> None:
    assert type(actual) is type(expected)
    if isinstance(expected, dict):
        assert isinstance(actual, dict)
        assert actual.keys() == expected.keys()
        for key, value in expected.items():
            _assert_strict_json(actual[key], value)
    elif isinstance(expected, list):
        assert isinstance(actual, list)
        assert len(actual) == len(expected)
        for actual_item, expected_item in zip(actual, expected, strict=True):
            _assert_strict_json(actual_item, expected_item)
    else:
        assert actual == expected


@pytest.mark.integration
@pytest.mark.asyncio
async def test_falsy_and_container_cells_survive_csv_authority_and_exports(
    tmp_path: Path,
    source_sidecar_binary: Path,
) -> None:
    cases = _load_cases()
    sidecar = Sidecar(
        source_sidecar_binary,
        _create_v2_workspace(tmp_path / "workspace"),
        workspace_identity={
            "VIBETABLE_WORKSPACE_ID": WORKSPACE_ID,
            "VIBETABLE_WORKSPACE_SESSION_EPOCH": str(SESSION_EPOCH),
            "VIBETABLE_WORKSPACE_FENCE_EPOCH": str(FENCE_EPOCH),
            "VIBETABLE_WORKSPACE_CLAIM_ID": CLAIM_ID,
        },
    )
    try:
        sidecar.start()
        table = _create_table(sidecar, "a5_falsy_roundtrip", "create-a5-falsy-table")
        fields: dict[str, str] = {}
        for case in cases:
            receipt = _create_field(
                sidecar,
                table,
                _field_draft(sidecar, table["tableId"], case),
                f"create-a5-{case['key']}",
            )
            fields[case["key"]] = _physical_name(receipt["definition"])

        source = tmp_path / "a5-falsy-source.csv"
        with source.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow([fields[case["key"]] for case in cases])
            writer.writerow([case["rawCell"] for case in cases])

        config = PocketBaseConfig(
            base_url=f"http://{sidecar.address}", session_secret=sidecar.secret
        )
        client = PocketBaseClient(
            transport=StdlibPocketBaseTransport(config), session_secret=sidecar.secret
        )
        tasks = build_task_service()
        runtime = ProductDataIoRuntime(client=client, task_service=tasks)
        grant = await tasks.register_host_import_source(
            HostImportSourceParams(
                path=str(source.resolve()),
                size_bytes=source.stat().st_size,
                mime_type="text/csv",
            )
        )
        plan = await runtime.preview_import(
            PreviewImportParams(
                grant_id=grant.grant_id,
                collection=table["tableId"],
                schema_revision=table["schemaRevision"],
            )
        )
        expected = {fields[case["key"]]: case["productValue"] for case in cases}
        assert plan.summary.total_rows == plan.summary.valid_rows == 1
        assert plan.summary.error_count == 0
        assert plan.unmatched_columns == []
        _assert_strict_json(plan.rows[0].values, expected)

        query = {"filters": [], "sorts": [], "offset": 0, "limit": 100}
        assert (await client.query_page(table_id=table["tableId"], query=query)).rows == []
        applied = await runtime.apply_import(
            ApplyImportParams(
                grant_id=grant.grant_id,
                collection=table["tableId"],
                token=plan.token.token,
                idempotency_prefix="a5-falsy-container",
            )
        )
        assert applied.created_count == 1
        assert applied.failed_rows == []
        authority_rows = (await client.query_page(table_id=table["tableId"], query=query)).rows
        assert len(authority_rows) == 1
        assert set(expected) <= authority_rows[0].keys()
        _assert_strict_json({field: authority_rows[0][field] for field in expected}, expected)

        exported: dict[str, dict[str, str]] = {}
        for export_format in ("csv", "xlsx"):
            target = tmp_path / f"a5-falsy-export.{export_format}"
            export_grant = await tasks.register_host_export_target(
                HostExportTargetParams(path=str(target.resolve()))
            )
            result = await runtime.export(
                ExportParams(
                    grant_id=export_grant.grant_id,
                    collection=table["tableId"],
                    query=query,
                    format=export_format,
                )
            )
            assert result.rows_written == 1
            if export_format == "csv":
                with target.open("r", encoding="utf-8-sig", newline="") as stream:
                    row = next(csv.DictReader(stream))
            else:
                workbook = load_workbook(target, read_only=True, data_only=False)
                try:
                    worksheet = workbook.active
                    assert worksheet is not None
                    rows = worksheet.iter_rows()
                    headers = [cell.value for cell in next(rows)]
                    cells = list(next(rows))
                    row = dict(zip(headers, (cell.value for cell in cells), strict=True))
                    formula_cell = cells[headers.index(fields["formula_text"])]
                    assert formula_cell.data_type == "s"
                finally:
                    workbook.close()
            exported[export_format] = {
                case["key"]: _export_text(row[fields[case["key"]]], case["logicalType"])
                for case in cases
            }

        expected_export = {case["key"]: case["exportText"] for case in cases}
        assert exported == {"csv": expected_export, "xlsx": expected_export}
    finally:
        sidecar.stop()

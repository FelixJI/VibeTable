from __future__ import annotations

import csv
import os
import subprocess
import tempfile
from collections.abc import Iterable, Iterator, Mapping
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from backend.adapters.pocketbase.client import PocketBaseClient
from backend.adapters.pocketbase.data_io import ProductDataIoRuntime
from backend.adapters.pocketbase.transport import PocketBaseConfig, StdlibPocketBaseTransport
from backend.application.task_service import build_task_service
from backend.contracts.data_io import ApplyImportParams, ExportParams, PreviewImportParams
from backend.contracts.task import HostExportTargetParams, HostImportSourceParams
from scripts.build_next import RepoPaths, build_sidecar_command
from tests.integration.packaged_sidecar_matrix import (
    CLAIM_ID,
    FENCE_EPOCH,
    SESSION_EPOCH,
    WORKSPACE_ID,
    Sidecar,
    _create_field,
    _create_table,
    _create_v2_workspace,
    _recommended_field_draft,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
EXPECTED_VALUES = {
    "nfc": "Caf\u00e9 \U0001f469\U0001f3fd\u200d\U0001f4bb",
    "nfd": "Cafe\u0301 \U0001f469\U0001f3fd\u200d\U0001f4bb",
}
EXPECTED_CODE_POINTS = {
    "nfc": (0x43, 0x61, 0x66, 0xE9, 0x20, 0x1F469, 0x1F3FD, 0x200D, 0x1F4BB),
    "nfd": (0x43, 0x61, 0x66, 0x65, 0x301, 0x20, 0x1F469, 0x1F3FD, 0x200D, 0x1F4BB),
}


def _build_source_sidecar(output_dir: Path) -> Path:
    output = output_dir / ("vibetable-pb.exe" if os.name == "nt" else "vibetable-pb")
    paths = RepoPaths.default(REPO_ROOT)
    subprocess.run(
        build_sidecar_command(
            paths,
            output=output,
            commit="unicode-data-io-roundtrip",
            build_time="2026-08-28T00:00:00Z",
        ),
        cwd=paths.sidecar_source_dir,
        check=True,
    )
    return output


@pytest.fixture
def source_sidecar_binary() -> Iterator[Path]:
    build_root = REPO_ROOT / "build" / "qa"
    build_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix="unicode-data-io-roundtrip-",
        dir=build_root,
    ) as run_root:
        yield _build_source_sidecar(Path(run_root))


def _physical_name(definition: Mapping[str, object]) -> str:
    identity = definition.get("identity")
    assert isinstance(identity, dict)
    physical_name = identity.get("physicalName")
    assert isinstance(physical_name, str)
    return physical_name


@pytest.mark.integration
@pytest.mark.asyncio
async def test_xlsx_native_dates_reach_go_authority_without_timezone_guessing(
    tmp_path: Path, source_sidecar_binary: Path
) -> None:
    sidecar = Sidecar(
        source_sidecar_binary,
        _create_v2_workspace(tmp_path / "workspace"),
        workspace_identity={
            "VIBETABLE_WORKSPACE_ID": WORKSPACE_ID,
            "VIBETABLE_WORKSPACE_SESSION_EPOCH": str(SESSION_EPOCH),
            "VIBETABLE_WORKSPACE_FENCE_EPOCH": str(FENCE_EPOCH),
            "VIBETABLE_WORKSPACE_CLAIM_ID": CLAIM_ID,
        },
    )
    try:
        sidecar.start()
        table = _create_table(sidecar, "native_dates", "create-native-dates")
        fields: list[str] = []
        for logical_type in ("date", "dateTime"):
            definition = _create_field(
                sidecar,
                table,
                _recommended_field_draft(sidecar, table["tableId"], logical_type, logical_type),
                f"create-native-{logical_type}",
            )["definition"]
            assert isinstance(definition, dict)
            fields.append(_physical_name(definition))
        date_field, datetime_field = fields
        source = tmp_path / "native-dates.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.append(fields)
        sheet.append([date(2026, 8, 29), datetime(2026, 8, 29, 14, 5, 6, 123000)])
        sheet.append(["1900-02-28", "2026-08-29T00:00:00+08:00"])
        workbook.save(source)
        workbook.close()

        config = PocketBaseConfig(
            base_url=f"http://{sidecar.address}", session_secret=sidecar.secret
        )
        client = PocketBaseClient(
            transport=StdlibPocketBaseTransport(config), session_secret=sidecar.secret
        )
        tasks = build_task_service()
        runtime = ProductDataIoRuntime(client=client, task_service=tasks)
        grant = await tasks.register_host_import_source(
            HostImportSourceParams(
                path=str(source.resolve()),
                size_bytes=source.stat().st_size,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )
        plan = await runtime.preview_import(
            PreviewImportParams(
                grant_id=grant.grant_id,
                collection=table["tableId"],
                schema_revision=table["schemaRevision"],
            )
        )
        assert plan.summary.total_rows == plan.summary.valid_rows == 2
        assert plan.summary.error_count == 0
        expected = {
            "2026-08-29": "2026-08-29T14:05:06.123Z",
            "1900-02-28": "2026-08-28T16:00:00Z",
        }
        assert {row.values[date_field]: row.values[datetime_field] for row in plan.rows} == expected
        query = {"filters": [], "sorts": [], "offset": 0, "limit": 100}
        assert (await client.query_page(table_id=table["tableId"], query=query)).rows == []
        applied = await runtime.apply_import(
            ApplyImportParams(
                grant_id=grant.grant_id,
                collection=table["tableId"],
                token=plan.token.token,
                idempotency_prefix="native-xlsx-dates",
            )
        )
        assert applied.created_count == 2
        assert applied.failed_rows == []
        page = await client.query_page(table_id=table["tableId"], query=query)
        # The query port exposes PocketBase's UTC date wire format, not preview DTOs.
        assert {row[date_field]: row[datetime_field] for row in page.rows} == {
            "2026-08-29 00:00:00.000Z": "2026-08-29 14:05:06.123Z",
            "1900-02-28 00:00:00.000Z": "2026-08-28 16:00:00.000Z",
        }
    finally:
        sidecar.stop()


def _assert_unicode_values(values: Mapping[str, str]) -> None:
    assert values == EXPECTED_VALUES
    assert {
        label: tuple(ord(character) for character in value) for label, value in values.items()
    } == EXPECTED_CODE_POINTS


def _labeled_values(
    rows: Iterable[Mapping[str, object]],
    *,
    label_field: str,
    value_field: str,
) -> dict[str, str]:
    values: dict[str, str] = {}
    for row in rows:
        label = row.get(label_field)
        value = row.get(value_field)
        assert isinstance(label, str)
        assert isinstance(value, str)
        values[label] = value
    return values


@pytest.mark.integration
@pytest.mark.asyncio
async def test_unicode_code_points_survive_import_authority_read_and_exports(
    tmp_path: Path,
    source_sidecar_binary: Path,
) -> None:
    data_dir = _create_v2_workspace(tmp_path / "workspace")
    sidecar = Sidecar(
        source_sidecar_binary,
        data_dir,
        workspace_identity={
            "VIBETABLE_WORKSPACE_ID": WORKSPACE_ID,
            "VIBETABLE_WORKSPACE_SESSION_EPOCH": str(SESSION_EPOCH),
            "VIBETABLE_WORKSPACE_FENCE_EPOCH": str(FENCE_EPOCH),
            "VIBETABLE_WORKSPACE_CLAIM_ID": CLAIM_ID,
        },
    )
    try:
        sidecar.start()
        table = _create_table(sidecar, "unicode_roundtrip", "create-unicode-roundtrip")
        label_definition = _create_field(
            sidecar,
            table,
            _recommended_field_draft(sidecar, table["tableId"], "label", "text"),
            "create-unicode-label",
        )["definition"]
        value_definition = _create_field(
            sidecar,
            table,
            _recommended_field_draft(sidecar, table["tableId"], "value", "text"),
            "create-unicode-value",
        )["definition"]
        assert isinstance(label_definition, dict)
        assert isinstance(value_definition, dict)
        label_field = _physical_name(label_definition)
        value_field = _physical_name(value_definition)

        source = tmp_path / "unicode-source.csv"
        with source.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow([label_field, value_field])
            writer.writerows(EXPECTED_VALUES.items())

        config = PocketBaseConfig(
            base_url=f"http://{sidecar.address}",
            session_secret=sidecar.secret,
        )
        client = PocketBaseClient(
            transport=StdlibPocketBaseTransport(config),
            session_secret=sidecar.secret,
        )
        tasks = build_task_service()
        runtime = ProductDataIoRuntime(client=client, task_service=tasks)
        import_grant = await tasks.register_host_import_source(
            HostImportSourceParams(
                path=str(source.resolve()),
                size_bytes=source.stat().st_size,
                mime_type="text/csv",
            )
        )
        plan = await runtime.preview_import(
            PreviewImportParams(
                grant_id=import_grant.grant_id,
                collection=table["tableId"],
                schema_revision=table["schemaRevision"],
                mode="create_only",
            )
        )
        assert plan.summary.total_rows == 2
        assert plan.summary.valid_rows == 2
        assert plan.summary.error_count == 0
        assert plan.unmatched_columns == []
        _assert_unicode_values(
            _labeled_values(
                (row.values for row in plan.rows),
                label_field=label_field,
                value_field=value_field,
            )
        )
        applied = await runtime.apply_import(
            ApplyImportParams(
                grant_id=import_grant.grant_id,
                collection=table["tableId"],
                token=plan.token.token,
                mode="create_only",
                idempotency_prefix="unicode-data-io-roundtrip",
            )
        )
        assert applied.created_count == 2
        assert applied.failed_rows == []

        query = {"filters": [], "sorts": [], "offset": 0, "limit": 100}
        authority_page = await client.query_page(table_id=table["tableId"], query=query)
        _assert_unicode_values(
            _labeled_values(
                authority_page.rows,
                label_field=label_field,
                value_field=value_field,
            )
        )

        csv_target = tmp_path / "unicode-export.csv"
        csv_grant = await tasks.register_host_export_target(
            HostExportTargetParams(path=str(csv_target.resolve()))
        )
        csv_result = await runtime.export(
            ExportParams(
                grant_id=csv_grant.grant_id,
                collection=table["tableId"],
                query=query,
                format="csv",
            )
        )
        assert csv_result.rows_written == 2
        with csv_target.open("r", encoding="utf-8-sig", newline="") as stream:
            csv_values = _labeled_values(
                csv.DictReader(stream),
                label_field=label_field,
                value_field=value_field,
            )
        _assert_unicode_values(csv_values)

        xlsx_target = tmp_path / "unicode-export.xlsx"
        xlsx_grant = await tasks.register_host_export_target(
            HostExportTargetParams(path=str(xlsx_target.resolve()))
        )
        xlsx_result = await runtime.export(
            ExportParams(
                grant_id=xlsx_grant.grant_id,
                collection=table["tableId"],
                query=query,
                format="xlsx",
            )
        )
        assert xlsx_result.rows_written == 2
        workbook = load_workbook(xlsx_target, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            assert worksheet is not None
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows)
            header_names: list[str] = []
            for header in headers:
                assert isinstance(header, str)
                header_names.append(header)
            xlsx_values = _labeled_values(
                (dict(zip(header_names, row, strict=True)) for row in rows),
                label_field=label_field,
                value_field=value_field,
            )
        finally:
            workbook.close()
        _assert_unicode_values(xlsx_values)
    finally:
        sidecar.stop()
